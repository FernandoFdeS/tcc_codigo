import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class GeraPlanilhaSaida:
    def __init__ (self, disciplinas,salas, horarios,x,caminho,nome_arquivo):
        self.disciplinas = disciplinas
        self.salas= salas
        self.horarios = horarios
        self.x = x 
        self.caminho = caminho
        self.nome_arquivo = nome_arquivo
        
    def cria_tabela_alocacoes(self,conflitos):
        alocacoes=[]
        for d in self.disciplinas:
            salas=[]
            for s in self.salas:
                for h in self.disciplinas[d].horarios_agrupamento():
                    if(round(self.x[d,s,h].X))==1:
                        if s not in salas:
                            salas.append(s)
            # Adiciona informações da disciplina "pai"
            alocacoes.append([
                d,
                self.disciplinas[d].nome_ccr,
                self.disciplinas[d].ch_ccr,
                self.disciplinas[d].nome_completo_curso,
                self.disciplinas[d].fase,
                self.disciplinas[d].horarioString[0],
                self.disciplinas[d].alunos,
                salas],
                )
            
            # Adiciona informações das disciplinas "filho", caso algum agrupamento tenha sido feito
            for agrupamentos in self.disciplinas[d].agrupamento:
                alocacoes.append([
                    agrupamentos.cod,
                    agrupamentos.nome_ccr,
                    agrupamentos.ch_ccr,
                    agrupamentos.nome_completo_curso,
                    agrupamentos.fase,
                    agrupamentos.horarioString[0],
                    agrupamentos.alunos,
                    salas],
                )
                        
        df1 = pd.DataFrame(alocacoes, columns=["cod", "nome_ccr", "ch_ccr", "curso", "fase", "horarios", "alunos", "sala"])
        df2 = pd.DataFrame(conflitos, columns=["SALA-TURNO","CODIGO-HORARIO"])

        nome_arquivo = self.caminho+"tabela_alocacoes.xlsx"
        with pd.ExcelWriter(nome_arquivo) as writer:
            df1.to_excel(writer, sheet_name='Alocações', index=False)
            df2.to_excel(writer, sheet_name='Conflitos', index=False)
    
    def exporta_alocacoes(self): 
        disciplinas_nao_alocadas=[]
        disciplinas_qtd_alocacoes = dict()
        
        # Coluna de horarios utilizado para criação da matriz usada para geração do aqruivo final
        coluna_horarios=["SEG-M","TER-M","QUA-M","QUI-M","SEX-M","SAB-M","#","SEG-V","TER-V","QUA-V","QUI-V","SEX-V","SAB-V","#","SEG-N","TER-N","QUA-N","QUI-N","SEX-N","SAB-N"]
        
        # Coluna de horarios utilizados como header no arquivo de saída para facilitar a compreensão
        coluna_horarios_csv=[["MATUTINO","-","-","-","-","-","#","VESPERTINO","-","-","-","-","-","#","NOTURNO","-","-","-","-","-"],
                        ["SEG","TER","QUA","QUI","SEX","SAB","#","SEG","TER","QUA","QUI","SEX","SAB","#","SEG","TER","QUA","QUI","SEX","SAB"]]

        linha_salas=[]
        blocoAtual = "A"
        linha_salas.append("BLOCO "+blocoAtual)
        for s in self.salas:
            blocoSala=s.split("-")[1]
            if blocoAtual!=blocoSala:
                blocoAtual=blocoSala
                linha_salas.append("BLOCO "+blocoAtual)
            linha_salas.append(s) 

        matriz = [['-' for coluna in range(len(coluna_horarios))] for linha in range(len(linha_salas))]
        for linha in linha_salas:
            if "BLOCO" in linha:
                index = linha_salas.index(linha)
                matriz[index] = ['#' for _ in range(len(coluna_horarios))]
        
        for coluna in range(len(coluna_horarios)):
            for linha in range(len(linha_salas)):
                if "#" in coluna_horarios[coluna]:
                    matriz[linha][coluna]="#"

        for d in self.disciplinas:
            disciplinas_qtd_alocacoes[d]=0
            for s in self.salas:
                for h in self.disciplinas[d].horarios:
                    if(round(self.x[d,s,h].X))==1:
                        disciplinas_qtd_alocacoes[d] += 1
            for agrupamentos in self.disciplinas[d].agrupamento:
                disciplinas_qtd_alocacoes[agrupamentos.cod]=0
                for s in self.salas:
                    for h in agrupamentos.horarios:
                        if(round(self.x[d,s,h].X))==1:
                            disciplinas_qtd_alocacoes[agrupamentos.cod] += 1

        # Preenche a matriz das alocações que sera utilizada no arquivo de saida
        for d in self.disciplinas:
            tem_alocacao_completa = True # A disciplina ou algum de seus agrupamentos teve todos os horarios alocados?
            if disciplinas_qtd_alocacoes[d] != len(self.disciplinas[d].horarios):
                tem_alocacao_completa = False
                disciplinas_nao_alocadas.append(self.disciplinas[d].formata_individual())

            if len(self.disciplinas[d].agrupamento) > 0 :
                for agrupamento in self.disciplinas[d].agrupamento:
                    if(disciplinas_qtd_alocacoes[agrupamento.cod] != len(agrupamento.horarios)):
                        disciplinas_nao_alocadas.append(agrupamento.formata_individual())
                    else:
                        tem_alocacao_completa = True
                        
            if(not tem_alocacao_completa):
               continue

            for s in self.salas:
                for h in self.disciplinas[d].horarios_agrupamento():
                    if(round(self.x[d,s,h].X))==1:
                        linha = linha_salas.index(s)
                        coluna=(coluna_horarios.index(self.horarios[h].converte_horario()))
                        if(matriz[linha][coluna]=='-'):
                            matriz[linha][coluna] = self.disciplinas[d].formata_saida(self.horarios[h].converte_horario())
                        elif (self.disciplinas[d].formata_saida(self.horarios[h].converte_horario()) not in matriz[linha][coluna]):
                            matriz[linha][coluna] = matriz[linha][coluna] +" | "+self.disciplinas[d].formata_saida(self.horarios[h].converte_horario())+" COMPARTILHAMENTO"         

        qtdNaoAlocadas=("Não Alocadas ("+str(len(disciplinas_nao_alocadas))+")")  

        # Verifica se todas as disciplinas foram alocadas
        if(len(disciplinas_nao_alocadas))==0:
            print("Alocações realizadas com sucesso!")
        else:
            print("Disciplinas alocadas: "+str((len(disciplinas_qtd_alocacoes)-len(disciplinas_nao_alocadas)))+"/"+str(len(disciplinas_qtd_alocacoes)))
            print("Disciplinas não alocadas/parcialmente alocadas: "+str(len(disciplinas_nao_alocadas)))
        print("Gerando planilha e tabela de alocações...")


        # Cria um DataFrame com rótulos personalizados
        while(len(disciplinas_nao_alocadas))<len(linha_salas):
            disciplinas_nao_alocadas.append("-")
        while(len(disciplinas_nao_alocadas))>len(linha_salas):
            linha_salas.append("-")
        indexes = pd.MultiIndex.from_arrays([disciplinas_nao_alocadas,linha_salas],names=[qtdNaoAlocadas,'Salas'])
        df = pd.DataFrame(matriz, columns=coluna_horarios_csv, index=indexes)        


        df.to_excel(self.caminho+self.nome_arquivo, index=True,engine='openpyxl')

        # Personalizando a planilha
        workbook = openpyxl.load_workbook(self.caminho+self.nome_arquivo)
        worksheet = workbook.worksheets[0]
        worksheet.merge_cells("C1:H1")
        worksheet.merge_cells("J1:O1")
        worksheet.merge_cells("Q1:V1")
        for coluna in range(ord("C"), ord("V") + 1):
            coluna_letra = chr(coluna)
            worksheet.column_dimensions[coluna_letra].width = (5 * 8.43)
        worksheet.column_dimensions["A"].width = (8.43*2.2)
        worksheet.column_dimensions["B"].width = (8.43*1.3)
        worksheet.column_dimensions["I"].width = (2)
        worksheet.column_dimensions["P"].width = (2)

        # Estilzando o HEADER (turnos)
        estilo_header = openpyxl.styles.NamedStyle(name="estilo_header")
        estilo_header.font = Font(name="Arial",color="000000", bold=True)  # Letra branca e negrito
        estilo_header.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Fundo preto
        estilo_header.alignment = Alignment(horizontal="center")  # Alinhamento centralizado
        estilo_header.border = Border(
            left=Side(style="thin", color="000000"),  # Borda esquerda branca
            right=Side(style="thin", color="000000"),  # Borda direita branca
            top=Side(style="thin", color="000000"),  # Borda superior branca
            bottom=Side(style="thin", color="000000"),  # Borda inferior branca
        )
        for coluna in worksheet.iter_cols(min_row=1, max_row=1):
            for celula in coluna:
                celula.style = estilo_header

        # Pintando os quadradinhos de cima de preto
        preenchimento_preto = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        ws = workbook.active
        celulas_pretas = ['A1', 'A2', 'B1', 'B2']
        for celula in celulas_pretas:
            ws[celula].fill = preenchimento_preto

        # Estilzando os indexs (Dias da semana & Salas) 
        estilo_index = openpyxl.styles.NamedStyle(name="estilo_index")
        estilo_index.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        estilo_index.font = Font(name="Arial", bold=True)
        estilo_index.alignment = Alignment(horizontal="center", vertical="center")
        estilo_index.border = Border(
            left=Side(style="thin", color="000000"),  # Borda esquerda branca
            right=Side(style="thin", color="000000"),  # Borda direita branca
            top=Side(style="thin", color="000000"),  # Borda superior branca
            bottom=Side(style="thin", color="000000"),  # Borda inferior branca
        )
        # Dias da semana
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=3, max_col=22):
            for cell in row:
                cell.style = estilo_index
        # Salas
        for row in ws.iter_rows(min_row=3, min_col=2,max_col=2):
            for cell in row:
                cell.style = estilo_index

        # Alocações
        for row in ws.iter_rows(min_row=3, min_col=3, max_col=22):  # Defina o intervalo C3:T3 até o final do arquivo
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Cor de células especiais
        fusao = PatternFill(start_color="ff7b59", end_color="ff7b59", fill_type="solid")
        fusao = openpyxl.styles.NamedStyle(name="fusao")
        fusao.fill = PatternFill(start_color="ff7b59", end_color="ff7b59", fill_type="solid")  # Fundo preto
        fusao.font = Font(name="Arial")
        fusao.alignment = Alignment(horizontal="center", vertical="center")
        fusao.border = Border(
            left=Side(style="thin", color="000000"),  # Borda esquerda branca
            right=Side(style="thin", color="000000"),  # Borda direita branca
            top=Side(style="thin", color="000000"),  # Borda superior branca
            bottom=Side(style="thin", color="000000"),  # Borda inferior branca
        )

        compartilhamento = PatternFill(start_color="ff7b59", end_color="ff7b59", fill_type="solid")
        compartilhamento = openpyxl.styles.NamedStyle(name="compartilhamento")
        compartilhamento.fill = PatternFill(start_color="ffa500", end_color="ff7b59", fill_type="solid")  # Fundo preto
        compartilhamento.font = Font(name="Arial")
        compartilhamento.alignment = Alignment(horizontal="center", vertical="center")
        compartilhamento.border = Border(
            left=Side(style="thin", color="000000"),  # Borda esquerda branca
            right=Side(style="thin", color="000000"),  # Borda direita branca
            top=Side(style="thin", color="000000"),  # Borda superior branca
            bottom=Side(style="thin", color="000000"),  # Borda inferior branca
        )

        agrupamento = PatternFill(start_color="ff7b59", end_color="ff7b59", fill_type="solid")
        agrupamento = openpyxl.styles.NamedStyle(name="agrupamento")
        agrupamento.fill = PatternFill(start_color="ab93f5", end_color="ab93f5", fill_type="solid")  # Fundo preto
        agrupamento.font = Font(name="Arial")
        agrupamento.alignment = Alignment(horizontal="center", vertical="center")
        agrupamento.border = Border(
            left=Side(style="thin", color="000000"),  # Borda esquerda branca
            right=Side(style="thin", color="000000"),  # Borda direita branca
            top=Side(style="thin", color="000000"),  # Borda superior branca
            bottom=Side(style="thin", color="000000"),  # Borda inferior branca
        )

        separador = PatternFill(start_color="ff7b59", end_color="ff7b59", fill_type="solid")
        separador = openpyxl.styles.NamedStyle(name="separador")
        separador.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Fundo preto
        separador.font = Font(name="Arial",color="FFFFFF",bold=True)
        separador.alignment = Alignment(horizontal="center", vertical="center")
        separador.border = Border(
            left=Side(style="thin", color="000000"),  # Borda esquerda branca
            right=Side(style="thin", color="000000"),  # Borda direita branca
            top=Side(style="thin", color="000000"),  # Borda superior branca
            bottom=Side(style="thin", color="000000"),  # Borda inferior branca
        )

        for row in ws.iter_rows(min_row=1, min_col=2):
            for cell in row:
                if cell.value is None:
                    cell.style = separador
                else:
                    if "FUSAO" in cell.value:
                        cell.style = fusao
                        cell.value = cell.value.replace("FUSAO","")
                    if "COMPARTILHAMENTO" in cell.value:
                        cell.style = compartilhamento
                        cell.value = cell.value.replace("COMPARTILHAMENTO","")
                    if "AGRUPAMENTO" in cell.value:
                        cell.style = agrupamento
                        cell.value = cell.value.replace("AGRUPAMENTO","")
                    if "BLOCO" in cell.value or "#" in cell.value:
                        cell.style = separador
                        cell.value = cell.value.replace("#","")


        workbook.save(self.caminho+self.nome_arquivo)

        workbook.close()
